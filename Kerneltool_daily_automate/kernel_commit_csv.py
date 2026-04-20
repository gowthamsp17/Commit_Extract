#!/usr/bin/env python3
"""
Linux Kernel Commit Details CSV Generator

Features:
  - Single-branch mode: accepts a pre-built stable-tree log file
    (format: "<version> <stable_hash> <description>" per line)
  - Mainline commit ID resolved from stable commit body (no git log --grep)
  - First Version taken directly from the log file (no git tag --contains)
  - Recursive Fixes: chain resolution
  - CONFIG_ parameter resolution per changed file
  - Applicability check via kernel .config (y-applicable / m-applicable / Not applicable)
  - Severity classification with reason
  - Subsystem tag (parsed from commit subject prefix)
  - Backport status summary column
  - Cc: stable / security tag detection
  - Output formats: csv (default), html, xlsx  via --format
  - Parallel processing via --parallel  (uses cpu_count // 2 workers)
  - Progress bar (tqdm if installed, plain fallback)
  - Summary report printed after CSV is written

Usage (new log-file mode — single branch):
    python kernel_commit_csv.py \\
        --repo   ~/linux \\
        --file   logs-6.1.y.txt \\
        --branch "6.1.y:origin/linux-6.1.y:config-6.1.123" \\
        --output results.csv [--format csv|html|xlsx] [--parallel]

Log file format (one entry per line):
    <version>  <stable_commit_hash>  <description>
    e.g.:  6.1.167  4ec349af3ef702d0d52cb2463bebef30739e65bd  selftests: net: ...

Prerequisites:
    - Run from within the Linux stable repo directory.
    - The remote for the configured branch must be fetched.
    - Optional: pip install tqdm openpyxl   (progress bar + xlsx output)
"""

import subprocess
import csv
import sys
import os
import re
import argparse
import concurrent.futures
import threading

# ── Configuration ─────────────────────────────────────────────────────────────

MAINLINE_BASE_URL = "https://github.com/torvalds/linux/commit"
STABLE_BASE_URL   = "https://git.kernel.org/pub/scm/linux/kernel/git/stable/linux.git/commit/"

# Single-branch mode: populated in main() from --branch.
# Structure: { label: {"ref": str, "cfg_path": str|None, "kernel_cfg": dict} }
# Exactly one entry is expected at runtime.
BRANCHES = {}

# Global map: stable_commit_hash → {"version": str, "mainline_id": str}
# Populated in main() by parsing the log file.  Used by process_commit() to
# avoid calling  git tag --contains  and  git log --grep  per commit.
STABLE_LOG_MAP: dict = {}

OUTPUT_DEFAULT = "kernel_commits.csv"

# Regex to detect "Fixes: abc1234def5 (...)" in commit bodies.
# Uses \r? to handle CRLF line endings that git can emit on some systems.
# Allows optional leading whitespace and is case-insensitive.
FIXES_RE = re.compile(
    r"(?:^|\r?\n)\s*Fixes:\s+([0-9a-f]{7,40})",
    re.IGNORECASE,
)


# ═══════════════════════════════════════════════════════════════════════════════
#  CONFIG_ resolution  (inlined from find_config.py)
# ═══════════════════════════════════════════════════════════════════════════════

def _read_makefile(path):
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            return f.readlines()
    return []


def _join_lines(lines):
    full, current = [], ""
    for line in lines:
        line = line.rstrip()
        if line.endswith("\\"):
            current += line[:-1] + " "
        else:
            current += line
            full.append(current)
            current = ""
    return full


def _find_direct_mapping(obj_name, lines):
    config_pat = re.compile(r'[A-Za-z0-9_\-]+-\$\((CONFIG_[A-Za-z0-9_]+)\)\s*[:+]?=\s*(.+)')
    built_pat  = re.compile(r'obj-y\s*[:+]?=\s*(.+)')
    mod_pat    = re.compile(r'obj-m\s*[:+]?=\s*(.+)')
    for line in lines:
        m = config_pat.search(line)
        if m and obj_name in m.group(2).split():
            return m.group(1)
        m = built_pat.search(line)
        if m and obj_name in m.group(1).split():
            return "BUILTIN"
        m = mod_pat.search(line)
        if m and obj_name in m.group(1).split():
            return "MODULE"
    return None


def _find_parent_object(obj_name, lines):
    pattern = re.compile(r'([A-Za-z0-9_\-]+)-(?:objs|y|m)\s*[:+]?=\s*(.+)')
    for line in lines:
        m = pattern.search(line)
        if m and obj_name in m.group(2).split():
            return m.group(1) + ".o"
    return None


def _resolve_object(obj_name, lines, visited=None):
    if visited is None:
        visited = set()
    if obj_name in visited:
        return None
    visited.add(obj_name)
    result = _find_direct_mapping(obj_name, lines)
    if result:
        return result
    parent = _find_parent_object(obj_name, lines)
    if parent:
        return _resolve_object(parent, lines, visited)
    return None


def _find_dir_mapping(dirname, lines):
    config_pat  = re.compile(r'obj-\$\((CONFIG_[A-Za-z0-9_]+)\)\s*[:+]?=\s*(.+)')
    built_pat   = re.compile(r'obj-y\s*[:+]?=\s*(.+)')
    mod_pat     = re.compile(r'obj-m\s*[:+]?=\s*(.+)')
    subdir_pat  = re.compile(r'subdir-\$\((CONFIG_[A-Za-z0-9_]+)\)\s*[:+]?=\s*(.+)')
    targets = [dirname + "/", dirname]
    for line in lines:
        m = config_pat.search(line)
        if m and any(t in m.group(2).split() for t in targets):
            return m.group(1)
        m = built_pat.search(line)
        if m and any(t in m.group(1).split() for t in targets):
            return "BUILTIN"
        m = mod_pat.search(line)
        if m and any(t in m.group(1).split() for t in targets):
            return "MODULE"
        m = subdir_pat.search(line)
        if m and any(t in m.group(2).split() for t in targets):
            return m.group(1)
    return None


def _find_hostprog(obj_base, lines):
    pattern = re.compile(r'hostprogs\s*[:+]?=\s*(.+)')
    for line in lines:
        m = pattern.search(line)
        if m and obj_base in m.group(1).split():
            return True
    return False


def _is_sentinel(result):
    return result in ("BUILTIN", "MODULE")


def _resolve_directory_config_recursive(directory):
    current_dir = os.path.normpath(directory)
    while True:
        parent_dir = os.path.dirname(current_dir)
        if parent_dir == current_dir:
            break
        dirname  = os.path.basename(current_dir)
        kbuild   = os.path.join(parent_dir, "Kbuild")
        makefile = os.path.join(parent_dir, "Makefile")
        lines, source_file = [], None
        if os.path.exists(kbuild):
            lines, source_file = _join_lines(_read_makefile(kbuild)), kbuild
        elif os.path.exists(makefile):
            lines, source_file = _join_lines(_read_makefile(makefile)), makefile
        if lines:
            result = _find_dir_mapping(dirname, lines)
            if result:
                if _is_sentinel(result):
                    current_dir = parent_dir
                    continue
                return result, source_file
        current_dir = parent_dir
    return None, None


def _resolve_source_file(file_path):
    file_path  = os.path.normpath(file_path)
    filename   = os.path.basename(file_path)
    directory  = os.path.dirname(file_path)
    obj_name   = re.sub(r'\.(c|rs)$', '.o', filename)
    base_name  = re.sub(r'\.(c|rs)$', '',   filename)
    search_dir = directory

    while True:
        kbuild   = os.path.join(search_dir, "Kbuild")
        makefile = os.path.join(search_dir, "Makefile")
        lines    = []
        if os.path.exists(kbuild):
            lines = _join_lines(_read_makefile(kbuild))
        elif os.path.exists(makefile):
            lines = _join_lines(_read_makefile(makefile))

        if lines:
            if _find_hostprog(base_name, lines):
                return None   # host program — skip

            result = _resolve_object(obj_name, lines)
            if result:
                if _is_sentinel(result):
                    config, _ = _resolve_directory_config_recursive(search_dir)
                    return config
                return result

            relative_obj = re.sub(r'\.(c|rs)$', '.o',
                                   os.path.relpath(file_path, search_dir))
            result = _resolve_object(relative_obj, lines)
            if result:
                if _is_sentinel(result):
                    config, _ = _resolve_directory_config_recursive(search_dir)
                    return config
                return result

            dirname = os.path.basename(search_dir)
            result  = _find_dir_mapping(dirname, lines)
            if result:
                if _is_sentinel(result):
                    config, _ = _resolve_directory_config_recursive(search_dir)
                    return config
                return result

        parent = os.path.dirname(search_dir)
        if parent == search_dir:
            break
        search_dir = parent

    return None


def _resolve_directory_config(directory):
    dirname    = os.path.basename(os.path.normpath(directory))
    search_dir = os.path.dirname(os.path.normpath(directory))
    while True:
        kbuild   = os.path.join(search_dir, "Kbuild")
        makefile = os.path.join(search_dir, "Makefile")
        lines, source_file = [], None
        if os.path.exists(kbuild):
            lines, source_file = _join_lines(_read_makefile(kbuild)), kbuild
        elif os.path.exists(makefile):
            lines, source_file = _join_lines(_read_makefile(makefile)), makefile
        if lines:
            result = _find_dir_mapping(dirname, lines)
            if result:
                if _is_sentinel(result):
                    config, src = _resolve_directory_config_recursive(search_dir)
                    return (config, src) if config else (None, None)
                return result, source_file
        parent = os.path.dirname(search_dir)
        if parent == search_dir:
            break
        dirname    = os.path.basename(search_dir)
        search_dir = parent
    return None, None


def resolve_config(file_path):
    """
    Return the CONFIG_ string for file_path, or None if unresolvable / not applicable.
    Only pure CONFIG_* strings are returned; all other messages become None.
    """
    if not os.path.exists(file_path):
        return None

    file_path = os.path.normpath(file_path)
    filename  = os.path.basename(file_path)
    directory = os.path.dirname(file_path)

    raw = None

    if filename.endswith(".c") or filename.endswith(".rs"):
        raw = _resolve_source_file(file_path)

    elif filename.endswith(".h"):
        c_counterpart = os.path.join(directory, filename[:-2] + ".c")
        if os.path.exists(c_counterpart):
            raw = _resolve_source_file(c_counterpart)

    elif filename in ("Makefile", "Kbuild", "Kconfig"):
        config, _ = _resolve_directory_config(directory)
        raw = config

    # Return only proper CONFIG_* tokens
    if raw and re.match(r'^CONFIG_[A-Za-z0-9_]+$', raw):
        return raw
    return None


# ── CONFIG_ resolution for a list of files ────────────────────────────────────

def get_config_params(files_str: str, repo_root: str) -> str:
    """
    Given a space-separated string of repo-relative file paths and the repo
    root, resolve CONFIG_ for each file.

    Returns a space-separated string of unique CONFIG_ values (preserving
    first-seen order).

    Fallback messages (when no CONFIG_ values found at all):
      "File not found"   — every file in the list is absent from the repo tree
      "No CONFIG_ found" — files exist but none resolved to a CONFIG_
    """
    if not files_str.strip():
        return ""

    files    = files_str.split()
    configs  = []
    seen     = set()
    n_found  = 0   # files that actually exist on disk
    n_total  = len(files)

    for f in files:
        abs_path = os.path.join(repo_root, f)
        if os.path.exists(abs_path):
            n_found += 1
            cfg = resolve_config(abs_path)
            if cfg and cfg not in seen:
                seen.add(cfg)
                configs.append(cfg)

    if configs:
        return " ".join(configs)

    # No CONFIG_ resolved — report why
    if n_found == 0:
        return "File not found"

    return "No CONFIG_ found"


# ═══════════════════════════════════════════════════════════════════════════════
#  Git helpers
# ═══════════════════════════════════════════════════════════════════════════════

def run(cmd: list) -> str:
    """Run a shell command and return stripped stdout. Returns '' on error."""
    try:
        result = subprocess.run(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
        return result.stdout.strip()
    except Exception as e:
        print(f"  [ERROR] Command failed: {' '.join(cmd)}\n  {e}")
        return ""


def get_full_sha(commit_id: str) -> str:
    return run(["git", "rev-parse", "--verify", commit_id + "^{commit}"])


def commit_exists(commit_id: str) -> bool:
    """Return True if commit_id is present in the local repo."""
    out = run(["git", "cat-file", "-t", commit_id])
    return out == "commit"


def ensure_commit(commit_id: str) -> bool:
    """
    Make sure commit_id is available locally.
    If not, attempt  git fetch origin <commit_id>  (works on GitHub-hosted
    repos and many others that allow single-object fetches).
    Returns True if the commit is available after the attempt.
    """
    if commit_exists(commit_id):
        return True

    print(f"    [fetch] {commit_id} not in local repo — trying git fetch origin {commit_id[:12]}...")
    run(["git", "fetch", "origin", commit_id])

    if commit_exists(commit_id):
        print(f"    [fetch] OK — commit now available.")
        return True

    print(f"    [fetch] FAILED — commit {commit_id[:12]} could not be fetched. "
          f"Run:  git fetch origin  and retry.")
    return False


def get_commit_description(commit_id: str) -> str:
    return run(["git", "log", "-1", "--format=%s", commit_id])


def get_commit_body(commit_id: str) -> str:
    return run(["git", "log", "-1", "--format=%B", commit_id])


def get_files_changed(commit_id: str) -> str:
    out = run(["git", "diff-tree", "--no-commit-id", "--name-only", "-r", commit_id])
    return " ".join(out.splitlines())


def get_commit_dates(commit_id: str) -> tuple:
    """
    Return (author_date, commit_date) for commit_id in YYYY-MM-DD HH:MM:SS format.

    Uses git's built-in --date=format: to produce the formatted string directly,
    avoiding any Python datetime parsing.

    Returns ("", "") on error.
    """
    out = run([
        "git", "show", "--no-patch",
        "--pretty=format:%ad%n%cd",
        "--date=format:%d/%m/%Y %H:%M:%S",
        commit_id,
    ])
    lines = out.splitlines()
    author_date = lines[0].strip() if len(lines) > 0 else ""
    commit_date = lines[1].strip() if len(lines) > 1 else ""
    return author_date, commit_date


def find_stable_commit(mainline_id: str, branch_ref: str) -> str:
    """
    Look up the stable commit hash for mainline_id from the pre-built
    STABLE_LOG_MAP (populated from the log file in main()).

    Falls back to the old  git log --grep  approach only when the map
    has no entry (e.g. when resolving Fix-N ancestors that were not in
    the input log).
    """
    # Fast path: consult the in-memory map built from the log file.
    entry = STABLE_LOG_MAP.get(mainline_id)
    if entry:
        return entry["stable_hash"]

    # Slow path (Fix-N ancestors not in the log file): git log --grep.
    out = run(["git", "log", branch_ref, f"--grep={mainline_id}", "--oneline"])
    if not out:
        return ""
    return out.splitlines()[0].split()[0]


def get_first_tag(commit_hash: str) -> str:
    """
    Return the first stable version that contains commit_hash.

    Primary source: STABLE_LOG_MAP, which maps stable_hash → version string
    taken directly from the input log file (e.g. "6.1.167").  This avoids
    the expensive  git tag --contains  call entirely for commits that came
    from the log file.

    Falls back to  git tag --contains  only for Fix-N ancestors that were
    not present in the log file.
    """
    if not commit_hash:
        return "N/A"

    # Fast path: version already known from the log file.
    entry = STABLE_LOG_MAP.get(commit_hash)
    if entry and entry.get("version"):
        return entry["version"]

    # Slow path: ask git (only for Fix-N commits not in the log).
    out = run(["git", "tag", "--contains", commit_hash])
    tags = [t.strip() for t in out.splitlines() if t.strip()]
    if not tags:
        return "N/A"
    try:
        from packaging.version import Version
        tags.sort(key=lambda t: Version(t.lstrip("v")))
    except Exception:
        def ver_key(t):
            nums = re.findall(r"\d+", t)
            return [int(n) for n in nums]
        tags.sort(key=ver_key)
    return tags[0]


def resolve_mainline_id(stable_hash: str) -> str:
    """
    Given a stable-tree commit hash, find the upstream (mainline) commit ID.

    Mirrors the validated bash script exactly:

        git show --no-patch --format=%B "$c" |
            grep -oE '[0-9a-f]{12,40}' |
            grep -v "^$c" |
            head -1

    That is:
      1. Get the full commit body as plain text.
      2. Extract every contiguous run of hex chars that is 12–40 chars long.
      3. Drop any token that matches the stable commit hash itself
         (full or abbreviated — bash matches the full hash via "^$c").
      4. Return the first surviving token.

    If no hex token survives (native stable commit, version bump, etc.) the
    stable commit IS the effective "mainline" reference — return stable_hash
    itself so the commit is still recorded rather than silently dropped.

    Returns '' only when the commit cannot be fetched locally at all.
    """
    if not ensure_commit(stable_hash):
        return ""

    body = get_commit_body(stable_hash)
    if not body:
        # Commit exists locally but has no message body.
        # Treat as native stable commit.
        return stable_hash

    # Replicate:  grep -oE '[0-9a-f]{12,40}' | grep -v "^$c" | head -1
    #
    # bash grep -oE prints each match on its own line, then grep -v "^$c"
    # drops lines whose entire content is the stable hash.  Because each
    # "line" here is exactly one regex match, "^$c" means the match == stable_hash.
    sha_re = re.compile(r'[0-9a-f]{12,40}')
    for candidate in sha_re.findall(body):
        if candidate == stable_hash:
            continue
        return candidate

    # No candidate found — native stable / version-bump commit.
    return stable_hash


def parse_log_file(log_path: str) -> dict:
    """
    Parse a stable-tree log file where each line is:
        <version>  <stable_hash>  <description...>

    Returns a dict keyed by full stable_hash (after git rev-parse):
        {
            stable_hash: {
                "version":     "6.1.167",
                "stable_hash": "<full 40-char sha>",
                "description": "...",
                "mainline_id": ""   # filled in later by resolve_mainline_id()
            },
            ...
        }

    Lines that are blank, start with '#', or don't have at least 3 fields
    are silently skipped.
    """
    entries = {}
    skipped = 0
    with open(log_path, "r", encoding="utf-8", errors="ignore") as fh:
        for lineno, raw in enumerate(fh, 1):
            line = raw.strip()
            if not line or line.startswith("#"):
                continue
            parts = line.split(None, 2)   # split on any whitespace, max 3 parts
            if len(parts) < 2:
                skipped += 1
                continue
            version     = parts[0]
            stable_hash = parts[1]
            description = parts[2] if len(parts) == 3 else ""
            entries[stable_hash] = {
                "version":     version,
                "stable_hash": stable_hash,
                "description": description,
                "mainline_id": "",
            }

    if skipped:
        print(f"[WARN] Skipped {skipped} malformed line(s) in log file.")
    print(f"[*] Log file parsed: {len(entries)} stable commits loaded.")
    return entries


def extract_fixes_sha(commit_id: str) -> str:
    """
    Parse the commit message for a 'Fixes: <sha>' tag.
    Returns the full resolved SHA of the fixed commit, or '' if none.
    Prints debug info so missing/undetected Fixes: tags are visible.
    """
    if not ensure_commit(commit_id):
        return ""

    body = get_commit_body(commit_id)

    if not body:
        print(f"    [WARN] Empty commit body for {commit_id[:12]} — cannot check for Fixes: tag.")
        return ""

    match = FIXES_RE.search(body)
    if not match:
        # Show the lines that contain "fixes" (case-insensitive) to help debug
        fixes_lines = [l.strip() for l in body.splitlines() if "fixes" in l.lower()]
        if fixes_lines:
            print(f"    [DEBUG] 'fixes' keyword found but regex did not match. Relevant lines:")
            for fl in fixes_lines:
                print(f"            {repr(fl)}")
        return ""

    short_sha = match.group(1)
    print(f"    Fixes: tag found → {short_sha}")
    full_sha = get_full_sha(short_sha)
    if not full_sha:
        print(f"    [WARN] Could not resolve short SHA '{short_sha}' — "
              f"try:  git fetch origin  to update your local repo.")
        # Still return the short SHA so the chain is not silently dropped
        return short_sha
    return full_sha


# ═══════════════════════════════════════════════════════════════════════════════
#  Applicability check
# ═══════════════════════════════════════════════════════════════════════════════



def load_kernel_config(kconfig_path: str) -> dict:
    """
    Parse a kernel .config file and return a dict of
    { 'CONFIG_FOO': 'y' | 'm' | 'n' | <value> }.
    Lines like  # CONFIG_FOO is not set  are stored as 'n'.
    """
    cfg = {}
    if not kconfig_path:
        return cfg
    path = os.path.abspath(os.path.expanduser(kconfig_path))
    if not os.path.exists(path):
        print(f"[WARN] Kernel config file not found: {path}")
        return cfg
    not_set_re = re.compile(r"^#\s+(CONFIG_[A-Za-z0-9_]+)\s+is not set")
    value_re   = re.compile(r"^(CONFIG_[A-Za-z0-9_]+)=(.+)$")
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            line = line.rstrip()
            m = not_set_re.match(line)
            if m:
                cfg[m.group(1)] = "n"
                continue
            m = value_re.match(line)
            if m:
                cfg[m.group(1)] = m.group(2).strip('"')
    print(f"[*] Loaded kernel config: {path}  ({len(cfg)} entries)")
    return cfg





def check_applicability(configs_str: str,
                        kernel_cfg: dict,
                        debug: bool = False) -> str:
    """
    Determine applicability for the commit based solely on the kernel .config.

    Priority order:
      1. Any CONFIG_ set to 'y' in kernel config  → "y-applicable"
      2. Any CONFIG_ set to 'm' in kernel config  → "m-applicable"
      3. All configs absent/n or no real CONFIG_  → "Not applicable"
      4. No kernel config supplied                → "" (blank)
    """
    # ── CONFIG_ check ─────────────────────────────────────────────────────────
    if not kernel_cfg:
        return ""   # no config supplied — leave blank

    # Extract only genuine CONFIG_* tokens (skip fallback messages)
    real_configs = [
        tok for tok in configs_str.split()
        if re.match(r'^CONFIG_[A-Za-z0-9_]+$', tok)
    ]

    if not real_configs:
        return "Not applicable"

    found_m = False
    for cfg_key in real_configs:
        val = kernel_cfg.get(cfg_key, "").lower()
        if debug:
            print(f"      [applicability] {cfg_key} = {val!r}")
        if val == "y":
            return "y-applicable"
        if val == "m":
            found_m = True

    if found_m:
        return "m-applicable"

    return "Not applicable"


# ═══════════════════════════════════════════════════════════════════════════════
#  Severity classification
# ═══════════════════════════════════════════════════════════════════════════════

# Each entry: (tier_label, reason_label, compiled_regex)
# Tiers are checked top-to-bottom; first match wins.
# reason_label is the human-readable text stored in "Severity Reason".

_SEVERITY_RULES = [

    # ── Critical ────────────────────────────────────────────────────────────
    ("Critical", "CVE reference",
        re.compile(r'\bCVE-\d{4}-\d+\b', re.IGNORECASE)),
    ("Critical", "Remote code execution",
        re.compile(r'\b(remote\s+code\s+execution|RCE)\b', re.IGNORECASE)),
    ("Critical", "Privilege escalation",
        re.compile(r'\b(privilege\s+escal|privesc)\b', re.IGNORECASE)),
    ("Critical", "Kernel panic",
        re.compile(r'\bkernel\s+panic\b', re.IGNORECASE)),
    ("Critical", "System crash",
        re.compile(r'\bsystem\s+crash\b', re.IGNORECASE)),
    ("Critical", "Use-after-free",
        re.compile(r'\buse.after.free\b', re.IGNORECASE)),
    ("Critical", "Double-free",
        re.compile(r'\bdouble.free\b', re.IGNORECASE)),
    ("Critical", "Buffer overflow",
        re.compile(r'\bbuffer\s+overflow\b', re.IGNORECASE)),
    ("Critical", "Heap overflow",
        re.compile(r'\bheap\s+overflow\b', re.IGNORECASE)),
    ("Critical", "Stack overflow",
        re.compile(r'\bstack\s+overflow\b', re.IGNORECASE)),
    ("Critical", "NULL pointer dereference",
        re.compile(r'\b(null[_\s\-]?pointer\s*deref|NULL\s+ptr\s+deref)\b', re.IGNORECASE)),
    ("Critical", "Out-of-bounds write",
        re.compile(r'\bout.of.bounds\s+write\b', re.IGNORECASE)),
    ("Critical", "Memory corruption",
        re.compile(r'\bmemory\s+corruption\b', re.IGNORECASE)),
    ("Critical", "Arbitrary code execution",
        re.compile(r'\barbitrary\s+(code|command)\s+execution\b', re.IGNORECASE)),
    ("Critical", "Security bypass",
        re.compile(r'\bsecurity\s+bypass\b', re.IGNORECASE)),
    ("Critical", "Crash (BUG/crash keyword)",
        re.compile(r'\bcrash\b', re.IGNORECASE)),

    # ── High ────────────────────────────────────────────────────────────────
    ("High", "Information/data leak",
        re.compile(r'\b(information\s+leak|info\s+leak|data\s+leak|infoleak)\b', re.IGNORECASE)),
    ("High", "Memory leak",
        re.compile(r'\bmemory\s+leak\b', re.IGNORECASE)),
    ("High", "Deadlock",
        re.compile(r'\bdeadlock\b', re.IGNORECASE)),
    ("High", "Race condition",
        re.compile(r'\brace\s+condition\b', re.IGNORECASE)),
    ("High", "Livelock",
        re.compile(r'\blivelock\b', re.IGNORECASE)),
    ("High", "Out-of-bounds read",
        re.compile(r'\bout.of.bounds\s+read\b', re.IGNORECASE)),
    ("High", "Out-of-bounds access",
        re.compile(r'\bout.of.bounds\b', re.IGNORECASE)),
    ("High", "Uninitialized memory/value",
        re.compile(r'\buninitialized\s+(memory|value|variable|data)\b', re.IGNORECASE)),
    ("High", "Integer overflow",
        re.compile(r'\binteger\s+overflow\b', re.IGNORECASE)),
    ("High", "Integer underflow",
        re.compile(r'\binteger\s+underflow\b', re.IGNORECASE)),
    ("High", "Type confusion",
        re.compile(r'\btype\s+confusion\b', re.IGNORECASE)),
    ("High", "Kernel Oops",
        re.compile(r'\bOops\b')),
    ("High", "BUG_ON triggered",
        re.compile(r'\bBUG_ON\b')),
    ("High", "WARN_ON triggered",
        re.compile(r'\bWARN_ON\b')),
    ("High", "Data corruption",
        re.compile(r'\bcorrupt(ion|ed|ing)?\b', re.IGNORECASE)),
    ("High", "Data loss",
        re.compile(r'\bdata\s+loss\b', re.IGNORECASE)),
    ("High", "Heap corruption",
        re.compile(r'\bheap\s+corrupt', re.IGNORECASE)),
    ("High", "Slab-out-of-bounds",
        re.compile(r'\bslab.out.of.bounds\b', re.IGNORECASE)),
    ("High", "Stack smashing",
        re.compile(r'\bstack\s+smash', re.IGNORECASE)),
    ("High", "UAF (abbreviation)",
        re.compile(r'\bUAF\b')),
    ("High", "OOB access",
        re.compile(r'\bOOB\b')),
    ("High", "Potential security issue",
        re.compile(r'\bsecurity\s+(issue|fix|flaw|vuln)', re.IGNORECASE)),
    ("High", "Vulnerability",
        re.compile(r'\bvulnerabilit', re.IGNORECASE)),

    # ── Medium ──────────────────────────────────────────────────────────────
    ("Medium", "Incorrect behavior/result",
        re.compile(r'\b(incorrect|wrong)\s+(behav|result|return|value|output|calculation|pte|mapping|address|order)\b', re.IGNORECASE)),
    ("Medium", "Regression",
        re.compile(r'\bregression\b', re.IGNORECASE)),
    ("Medium", "Performance regression/degradation",
        re.compile(r'\bperformance\s+(regression|degradation|issue|drop)\b', re.IGNORECASE)),
    ("Medium", "Suspend/resume issue",
        re.compile(r'\b(suspend|resume)\b', re.IGNORECASE)),
    ("Medium", "Boot failure/hang",
        re.compile(r'\bboot\s+(fail|loop|hang|crash|problem)\b', re.IGNORECASE)),
    ("Medium", "System hang",
        re.compile(r'\bhang\b', re.IGNORECASE)),
    ("Medium", "System freeze",
        re.compile(r'\bfreez(e|ing|es)\b', re.IGNORECASE)),
    ("Medium", "Timeout",
        re.compile(r'\btimeout\b', re.IGNORECASE)),
    ("Medium", "Spurious behavior",
        re.compile(r'\bspurious\b', re.IGNORECASE)),
    ("Medium", "Missing lock/unlock/check/barrier",
        re.compile(r'\bmissing\s+(lock|unlock|check|barrier|ref|put|get)\b', re.IGNORECASE)),
    ("Medium", "Divide-by-zero",
        re.compile(r'\bdivide.by.zero\b', re.IGNORECASE)),
    ("Medium", "kmalloc failure",
        re.compile(r'\bkmalloc\s+fail', re.IGNORECASE)),
    ("Medium", "NULL dereference (fix)",
        re.compile(r'\bnull\s+deref', re.IGNORECASE)),
    ("Medium", "Fix incorrect restore/pte/mapping",
        re.compile(r'\bfix\s+(incorrect|wrong|bad|broken)\b', re.IGNORECASE)),
    ("Medium", "Panic (non-critical path)",
        re.compile(r'\bpanic\b', re.IGNORECASE)),
    ("Medium", "Lost/missing wakeup",
        re.compile(r'\b(lost|missing)\s+wakeup\b', re.IGNORECASE)),
    ("Medium", "Soft lockup",
        re.compile(r'\bsoft\s+lockup\b', re.IGNORECASE)),
    ("Medium", "Hard lockup",
        re.compile(r'\bhard\s+lockup\b', re.IGNORECASE)),
    ("Medium", "RCU stall",
        re.compile(r'\bRCU\s+stall\b', re.IGNORECASE)),
    ("Medium", "NULL check missing",
        re.compile(r'\bmissing\s+null\s+check\b', re.IGNORECASE)),
    ("Medium", "Error path fix",
        re.compile(r'\berror\s+(path|handling|return|check)\b', re.IGNORECASE)),
    ("Medium", "Resource leak",
        re.compile(r'\bresource\s+leak\b', re.IGNORECASE)),
    ("Medium", "Fix (general bug fix)",
        re.compile(r'\bfix\b', re.IGNORECASE)),

    # ── Low ─────────────────────────────────────────────────────────────────
    ("Low", "Typo/spelling/grammar",
        re.compile(r'\b(typo|spelling|grammar)\b', re.IGNORECASE)),
    ("Low", "Whitespace fix",
        re.compile(r'\bwhitespace\b', re.IGNORECASE)),
    ("Low", "Indentation fix",
        re.compile(r'\bindent(ation)?\b', re.IGNORECASE)),
    ("Low", "Comment fix",
        re.compile(r'\bcomment\s+(fix|update|correct)\b', re.IGNORECASE)),
    ("Low", "Code cleanup",
        re.compile(r'\bcleanup\b', re.IGNORECASE)),
    ("Low", "Refactor",
        re.compile(r'\brefactor\b', re.IGNORECASE)),
    ("Low", "Code style",
        re.compile(r'\bcode\s+style\b', re.IGNORECASE)),
    ("Low", "Kconfig fix/cleanup",
        re.compile(r'\bkconfig\s+(fix|cleanup|update)\b', re.IGNORECASE)),
    ("Low", "Documentation fix/update",
        re.compile(r'\bdocument(ation)?\s+(fix|update|add|correct)\b', re.IGNORECASE)),
    ("Low", "printk/printf format fix",
        re.compile(r'\bprint[kf]?\s+(fix|format|correct)\b', re.IGNORECASE)),
    ("Low", "Compiler warning fix",
        re.compile(r'\bwarning\s+(fix|suppress|remov)\b', re.IGNORECASE)),
    ("Low", "Unused variable/function",
        re.compile(r'\bunused\s+(variable|import|function|header|parameter)\b', re.IGNORECASE)),
    ("Low", "Dead code removal",
        re.compile(r'\bdead\s+code\b', re.IGNORECASE)),
    ("Low", "Add missing header/include",
        re.compile(r'\b(add|missing)\s+(header|include)\b', re.IGNORECASE)),
    ("Low", "Endian fix",
        re.compile(r'\bendian\b', re.IGNORECASE)),
    ("Low", "Sparse warning fix",
        re.compile(r'\bsparse\b', re.IGNORECASE)),
    ("Low", "Checkpatch fix",
        re.compile(r'\bcheckpatch\b', re.IGNORECASE)),
    ("Low", "Rename/move",
        re.compile(r'\b(rename|move)\s+(file|symbol|function|variable|struct)\b', re.IGNORECASE)),
]


def classify_severity(commit_id: str):
    """
    Classify commit severity by scanning the full commit message body.

    Priority:
      0. Fixes: tag present                        → High / "Fixes tag present"
      1–N. _SEVERITY_RULES  (Critical → Low)

    Returns (severity_label, reason_label).
    severity_label: 'Critical' | 'High' | 'Medium' | 'Low' | 'Unknown'
    reason_label  : matched keyword description, or '' if Unknown.
    """
    body = get_commit_body(commit_id)
    if not body:
        return "Unknown", ""

    # ── Step 0: Fixes: tag takes immediate precedence ─────────────────────────
    if FIXES_RE.search(body):
        return "High", "Fixes tag present"

    # ── Steps 1-N: keyword rules ──────────────────────────────────────────────
    for tier_label, reason_label, pat in _SEVERITY_RULES:
        if pat.search(body):
            return tier_label, reason_label

    return "Unknown", ""


# ═══════════════════════════════════════════════════════════════════════════════
#  Subsystem tag
# ═══════════════════════════════════════════════════════════════════════════════

def get_subsystem(description: str) -> str:
    """
    Parse the kernel commit subject line for its real subsystem prefix.

    Linux subjects follow the convention:
        subsystem: short description
        subsystem: component: short description

    Some top-level prefixes are NOT subsystems — they describe the type of
    change (e.g. 'selftests', 'Documentation', 'tools') rather than the
    kernel component being fixed.  When one of these generic prefixes is
    found, the function looks at the NEXT colon-separated token instead.

    Examples:
        'mm: fix folio leak'                       → 'mm'
        'net/ipv4: fix checksum'                   → 'net/ipv4'
        'Bluetooth: hci_uart: fix ...'             → 'Bluetooth'
        'selftests: mptcp: pm: ensure ...'         → 'mptcp'
        'selftests/mptcp: pm: ensure ...'          → 'mptcp'
        'Documentation: admin-guide: update'       → 'admin-guide'
        'tools: perf: fix ...'                     → 'perf'
        'drm/i915: fix use-after-free'             → 'drm/i915'
        'MAINTAINERS: update email'                → ''
    """
    _GENERIC_PREFIXES = {
        "selftests", "selftest",
        "documentation", "doc",
        "tools",
        "samples",
        "scripts",
        "maintainers",
        "licenses",
        "headers",
        "uapi",
    }

    if not description:
        return ""

    parts = description.split(":")

    for part in parts:
        token = part.strip()
        if not token or " " in token:
            break

        # For generic-prefix detection only, check the first slash-component.
        # e.g. "selftests/mptcp" → first component is "selftests" (generic),
        # but "net/ipv4" → first component is "net" (not generic → keep full token).
        first_component = token.split("/")[0].strip().lower()

        if first_component in _GENERIC_PREFIXES:
            # The token might be "selftests/mptcp" — strip the generic prefix
            # and return the remainder (e.g. "mptcp") if it exists,
            # otherwise move on to the next colon-token.
            slash_parts = token.split("/", 1)
            if len(slash_parts) > 1 and slash_parts[1].strip():
                remainder = slash_parts[1].strip()
                if len(remainder) <= 40 and " " not in remainder:
                    return remainder
            continue

        if len(token) > 40:
            break

        return token   # first non-generic, non-prose token is the subsystem

    return ""


# ═══════════════════════════════════════════════════════════════════════════════
#  Backport status
# ═══════════════════════════════════════════════════════════════════════════════

def get_backport_status(info: dict) -> str:
    """
    Compute a human-readable backport summary from the branch columns already
    collected in `info`.
    """
    found   = []
    missing = []
    for label in BRANCHES:
        if info.get(f"{label} Commit Hash", "N/A") != "N/A":
            found.append(label)
        else:
            missing.append(label)


    if not found:
        return "No branches"
    if not missing:
        return "All branches"
    return f"Found: {', '.join(found)} | Missing: {', '.join(missing)}"


# ── Cc tag detection ──────────────────────────────────────────────────────────

_CC_STABLE_RE   = re.compile(r'^\s*Cc\s*:.*stable@vger\.kernel\.org',   re.IGNORECASE | re.MULTILINE)
_CC_SECURITY_RE = re.compile(r'^\s*Cc\s*:.*security@kernel\.org',        re.IGNORECASE | re.MULTILINE)


def check_cc_tags(commit_body: str) -> tuple:
    """
    Scan the commit message for Cc: stable and Cc: security tags.
    Returns (cc_stable: str, cc_security: str) each being 'Yes' or 'No'.
    """
    cc_stable   = "Yes" if _CC_STABLE_RE.search(commit_body)   else "No"
    cc_security = "Yes" if _CC_SECURITY_RE.search(commit_body) else "No"
    return cc_stable, cc_security



# ═══════════════════════════════════════════════════════════════════════════════
#  Progress bar  (tqdm if available, plain counter fallback)
# ═══════════════════════════════════════════════════════════════════════════════

try:
    from tqdm import tqdm as _tqdm_cls
    _TQDM_AVAILABLE = True
except ImportError:
    _TQDM_AVAILABLE = False


class _PlainProgress:
    """Minimal tqdm-compatible wrapper using plain print."""
    def __init__(self, total, desc=""):
        self.total   = total
        self.current = 0
        self.desc    = desc

    def update(self, n=1):
        self.current += n
        pct = int(100 * self.current / self.total) if self.total else 0
        print(f"  [{self.desc}] {self.current}/{self.total}  ({pct}%)", flush=True)

    def close(self):
        pass

    def __enter__(self):
        return self

    def __exit__(self, *_):
        self.close()


def make_progress(total: int, desc: str = "Processing"):
    if _TQDM_AVAILABLE:
        return _tqdm_cls(total=total, desc=desc, unit="commit",
                         dynamic_ncols=True, leave=True)
    return _PlainProgress(total=total, desc=desc)


# ═══════════════════════════════════════════════════════════════════════════════
#  Summary report
# ═══════════════════════════════════════════════════════════════════════════════

def print_summary(rows: list) -> None:
    """Print a summary table to stdout after all commits are processed."""

    def counter(rows, key):
        counts = {}
        for row in rows:
            val = row.get(key, "") or "Unknown"
            counts[val] = counts.get(val, 0) + 1
        return counts

    severity_order     = ["Critical", "High", "Medium", "Low", "Unknown"]
    applicability_order = [
        "y-applicable", "m-applicable", "Not applicable", "",
    ]
    backport_order = ["All branches", "No branches"]

    sep = "─" * 42

    print(f"\n{'═' * 42}")
    print("  SUMMARY REPORT")
    print(f"{'═' * 42}")
    print(f"  Total commits processed : {len(rows)}")

    # ── Severity ─────────────────────────────────────────────────────────────
    print(f"\n  Severity{' ' * 15}Count")
    print(f"  {sep}")
    sev_counts = counter(rows, "Severity")
    for label in severity_order:
        n = sev_counts.get(label, 0)
        if n:
            print(f"  {label:<24} {n:>5}")

    # ── Applicability ─────────────────────────────────────────────────────────
    # Key is "<label> Applicability" — derive from BRANCHES (single branch).
    app_label = f"{next(iter(BRANCHES))} Applicability" if BRANCHES else "Applicability"
    print(f"\n  Applicability{' ' * 10}Count")
    print(f"  {sep}")
    app_counts = counter(rows, app_label)
    for label in applicability_order:
        n = app_counts.get(label, 0)
        if n:
            display = label if label else "(blank — no config)"
            print(f"  {display:<24} {n:>5}")

    # ── Backport status ───────────────────────────────────────────────────────
    print(f"\n  Backport Status{' ' * 8}Count")
    print(f"  {sep}")
    bp_counts = counter(rows, "Backport Status")
    for label, n in sorted(bp_counts.items(), key=lambda x: -x[1]):
        print(f"  {label:<24} {n:>5}")

    # ── Subsystem top-10 ─────────────────────────────────────────────────────
    print(f"\n  Top Subsystems{' ' * 9}Count")
    print(f"  {sep}")
    sub_counts = counter(rows, "Subsystem")
    sub_counts.pop("", None)
    sub_counts.pop("Unknown", None)
    for label, n in sorted(sub_counts.items(), key=lambda x: -x[1])[:10]:
        print(f"  {label:<24} {n:>5}")

    print(f"{'═' * 42}\n")


def collect_commit_info(commit_id: str, repo_root: str,
                        is_fix_level: bool = False) -> dict:
    """
    Return a flat dict with all details for ONE stable-tree commit.

    commit_id     — the stable commit hash (from the log file or Fixes: chain).
    is_fix_level  — True when called for Fix-1/2/… ancestors.
                    Skips First Version only (not in the input log).
                    AuthorDate and CommitDate are collected at all levels.

    Columns produced (base level):
        <label> Commit Hash
        <label> GitHub Link
        <label> First Version       ← from STABLE_LOG_MAP (free); base only
        <label> AuthorDate          ← git show --pretty=fuller
        <label> CommitDate          ← git show --pretty=fuller
        Description
        Subsystem
        Files Changed
        CONFIG_ Parameters

    Fix-N level omits: First Version only
    """
    if not ensure_commit(commit_id):
        print(f"    [WARN] Skipping detail collection — commit not available locally.")
        stub = {
            "Description"        : "commit not found locally",
            "Subsystem"          : "",
            "Files Changed"      : "",
            "CONFIG_ Parameters" : "",
        }
        for label in BRANCHES:
            stub[f"{label} Commit Hash"]   = commit_id
            stub[f"{label} GitHub Link"]   = f"{STABLE_BASE_URL}?id={commit_id}"
            if not is_fix_level:
                stub[f"{label} First Version"] = "N/A"
            stub[f"{label} AuthorDate"]    = ""
            stub[f"{label} CommitDate"]    = ""
            stub[f"{label} Applicability"] = ""
        return stub

    description = get_commit_description(commit_id)
    files       = get_files_changed(commit_id)
    subsystem   = get_subsystem(description)

    print(f"    Description : {description}")
    print(f"    Subsystem   : {subsystem if subsystem else '(none)'}")
    print(f"    Files       : {files}")

    configs = get_config_params(files, repo_root)
    print(f"    CONFIG_     : {configs}")

    info = {}

    # ── Per-branch stable commit columns ─────────────────────────────────────
    for label, meta in BRANCHES.items():
        # The stable hash IS commit_id (this function is now called with the
        # stable hash directly).  Build the GitHub link from STABLE_BASE_URL.
        stable_url = f"{STABLE_BASE_URL}?id={commit_id}"
        info[f"{label} Commit Hash"] = commit_id
        info[f"{label} GitHub Link"] = stable_url

        if not is_fix_level:
            # First Version: free from STABLE_LOG_MAP for base-level commits.
            log_entry = STABLE_LOG_MAP.get(commit_id)
            if log_entry:
                first_ver = log_entry["version"]
                is_native = (log_entry["stable_hash"] == commit_id)
                tag_note  = " [native]" if is_native else ""
                print(f"    [{label}] version (from log): {first_ver}{tag_note}")
            else:
                # Fix-chain ancestor that happened to also be in the log —
                # fall back to git tag (rare).
                first_ver = get_first_tag(commit_id)
                print(f"    [{label}] version (git tag): {first_ver}")
            info[f"{label} First Version"] = first_ver

        # AuthorDate / CommitDate — collected at base and Fix-N levels
        author_date, commit_date = get_commit_dates(commit_id)
        info[f"{label} AuthorDate"] = author_date
        info[f"{label} CommitDate"] = commit_date
        print(f"    AuthorDate  : {author_date}")
        print(f"    CommitDate  : {commit_date}")

    # ── Content columns (shared across base and Fix-N) ────────────────────────
    info["Description"]        = description
    info["Subsystem"]          = subsystem
    info["Files Changed"]      = files
    info["CONFIG_ Parameters"] = configs

    return info


def prefix_dict(d: dict, prefix: str) -> dict:
    return {f"{prefix} {k}": v for k, v in d.items()}


def process_commit(commit_id: str, repo_root: str,
                   debug: bool = False,
                   stable_hash: str = "") -> dict:
    """
    Process one stable-tree commit and recursively follow its Fixes: chain.

    commit_id   — the stable commit hash (same as stable_hash for the primary
                  commit; kept as parameter name for compatibility with the
                  parallel worker).
    stable_hash — same as commit_id for primary commits (passed from main).

    All processing (description, dates, CONFIG_, severity, Cc tags) is done
    directly against the stable commit hash.  There is no mainline lookup.

    Applicability is determined solely from the kernel .config supplied via
    --branch.  No file/directory exclusion lists are used.

    Fix-N ancestors: we follow the Fixes: tag in the stable commit body.
    Those ancestor commits are also stable-tree hashes.  For Fix-N levels
    we skip First Version, AuthorDate, CommitDate (not in the log file and
    expensive to compute).

    Returns a single flat dict:
      (no prefix)  – the primary commit  [with Applicability, Severity, etc.]
      Fix-1        – the commit it directly fixes  [content cols only]
      Fix-2        – what Fix-1 fixes
      ...
    """
    # For primary commits stable_hash == commit_id (both passed from main).
    # Use commit_id throughout so the function signature stays compatible
    # with _process_one.
    print(f"\n[*] Processing stable commit: {commit_id[:12]}")

    ensure_commit(commit_id)

    row     = {}
    visited = set()

    info = collect_commit_info(commit_id, repo_root, is_fix_level=False)

    # ── Per-branch Applicability — only for the primary commit ───────────────
    files_str   = info.get("Files Changed", "")
    configs_str = info.get("CONFIG_ Parameters", "")
    for label, meta in BRANCHES.items():
        app = check_applicability(
            configs_str,
            meta["kernel_cfg"],
            debug=debug,
        )
        info[f"{label} Applicability"] = app
        print(f"    [{label}] Applicability: {app if app else '(no config supplied)'}")

    # ── Severity ──────────────────────────────────────────────────────────────
    severity, severity_reason = classify_severity(commit_id)
    info["Severity"]        = severity
    info["Severity Reason"] = severity_reason
    print(f"    Severity     : {severity} ({severity_reason if severity_reason else 'no match'})")

    # ── Backport Status ───────────────────────────────────────────────────────
    backport_status = get_backport_status(info)
    info["Backport Status"] = backport_status
    print(f"    Backport     : {backport_status}")

    # ── Cc: tags ──────────────────────────────────────────────────────────────
    commit_body = get_commit_body(commit_id)
    cc_stable, cc_security = check_cc_tags(commit_body)
    info["Cc stable"]   = cc_stable
    info["Cc security"] = cc_security
    print(f"    Cc stable    : {cc_stable}  |  Cc security: {cc_security}")

    row.update(info)
    visited.add(commit_id)

    current_id = commit_id
    depth      = 1

    while True:
        fixes_sha = extract_fixes_sha(current_id)

        if not fixes_sha:
            print(f"    No 'Fixes:' tag in {current_id[:12]} — chain ends here.")
            break

        if fixes_sha in visited:
            print(f"    Fixes: {fixes_sha[:12]} already processed — stopping to avoid loop.")
            break

        print(f"\n  [Fix-{depth}] {current_id[:12]}  →  fixes {fixes_sha[:12]}")
        visited.add(fixes_sha)

        # Fix-N: skip First Version / AuthorDate / CommitDate
        fix_info = collect_commit_info(fixes_sha, repo_root, is_fix_level=True)
        row.update(prefix_dict(fix_info, f"Fix-{depth}"))

        current_id = fixes_sha
        depth     += 1

    return row


# ═══════════════════════════════════════════════════════════════════════════════
#  Dynamic fieldname builder
# ═══════════════════════════════════════════════════════════════════════════════

# Content columns that appear at EVERY level (base + Fix-N).
# Branch-specific columns are inserted before these by build_fieldnames().
CONTENT_COLUMNS = [
    "Description",
    "Subsystem",
    "Files Changed",
    "CONFIG_ Parameters",
]

# Extra columns that appear ONLY at the base (unprefixed) level.
BASE_ONLY_COLUMNS = [
    "Applicability",      # placeholder — actual key is "<label> Applicability"
    "Severity",
    "Severity Reason",
    "Backport Status",
    "Cc stable",
    "Cc security",
]

# Branch columns produced at BASE level (in order):
#   <label> Commit Hash
#   <label> GitHub Link
#   <label> First Version
#   <label> AuthorDate
#   <label> CommitDate
#   <label> Applicability      ← base only
#
# Branch columns produced at FIX-N level (in order):
#   <label> Commit Hash
#   <label> GitHub Link
#   (First Version, AuthorDate, CommitDate omitted — not in log, expensive)


def _all_branch_labels(rows: list) -> list:
    """
    Return a sorted union of branch labels from the current BRANCHES dict
    and any labels already present in existing rows (detected by columns
    ending in ' Commit Hash' without a 'Fix-N ' prefix).
    Sorted by version number so columns are always in ascending order.
    """
    def _ver_key(lbl):
        return [int(n) for n in re.findall(r"\d+", lbl)]

    labels = set(BRANCHES.keys())
    for row in rows:
        for key in row:
            if key.endswith(" Commit Hash") and not re.match(r"^Fix-\d+ ", key):
                labels.add(key[: -len(" Commit Hash")])

    return sorted(labels, key=_ver_key)


def build_fieldnames(rows: list) -> list:
    """
    Build the fully-ordered column list that matches the requested layout:

    Base level:
        <label> Commit Hash
        <label> GitHub Link
        <label> First Version
        <label> AuthorDate
        <label> CommitDate
        Description
        Subsystem
        Files Changed
        CONFIG_ Parameters
        <label> Applicability
        Severity
        Severity Reason
        Backport Status
        Cc stable
        Cc security

    Fix-N level (no First Version):
        Fix-N <label> Commit Hash
        Fix-N <label> GitHub Link
        Fix-N <label> AuthorDate
        Fix-N <label> CommitDate
        Fix-N Description
        Fix-N Subsystem
        Fix-N Files Changed
        Fix-N CONFIG_ Parameters
    """
    max_depth = 0
    for row in rows:
        for key in row:
            m = re.match(r"^Fix-(\d+) ", key)
            if m:
                max_depth = max(max_depth, int(m.group(1)))

    sorted_labels = _all_branch_labels(rows)

    # ── Base level columns ────────────────────────────────────────────────────
    base_fieldnames = []
    for label in sorted_labels:
        base_fieldnames += [
            f"{label} Commit Hash",
            f"{label} GitHub Link",
            f"{label} First Version",
            f"{label} AuthorDate",
            f"{label} CommitDate",
        ]
    base_fieldnames += CONTENT_COLUMNS
    for label in sorted_labels:
        base_fieldnames.append(f"{label} Applicability")
    base_fieldnames += [
        "Severity",
        "Severity Reason",
        "Backport Status",
        "Cc stable",
        "Cc security",
    ]

    # ── Fix-N level columns ───────────────────────────────────────────────────
    fix_fieldnames = []
    for depth in range(1, max_depth + 1):
        pfx = f"Fix-{depth} "
        for label in sorted_labels:
            fix_fieldnames += [
                f"{pfx}{label} Commit Hash",
                f"{pfx}{label} GitHub Link",
                f"{pfx}{label} AuthorDate",
                f"{pfx}{label} CommitDate",
            ]
        fix_fieldnames += [f"{pfx}{col}" for col in CONTENT_COLUMNS]

    return base_fieldnames + fix_fieldnames


# ═══════════════════════════════════════════════════════════════════════════════
#  Output writers  (CSV · HTML · XLSX)
# ═══════════════════════════════════════════════════════════════════════════════

# Severity → background colour used in HTML and XLSX
_SEVERITY_COLOURS = {
    "Critical": "#ff4d4d",   # red
    "High":     "#ff944d",   # orange
    "Medium":   "#ffd24d",   # yellow
    "Low":      "#b3e6b3",   # light green
    "Unknown":  "#e0e0e0",   # grey
}

# Applicability → background colour
_APPLICABILITY_COLOURS = {
    "y-applicable":   "#99ff99",   # green
    "m-applicable":   "#ccffcc",   # light green
    "Not applicable": "#ffcccc",   # light red
}


def load_existing_commit_ids(output_path: str, fmt: str) -> set:
    """
    Read the existing output file (if it exists) and return a set of all
    stable commit hashes already recorded (from the '<label> Commit Hash' column).

    Supports CSV and XLSX; HTML is not parsed (returns empty set with a warning).
    Returns an empty set when the file does not exist yet.
    """
    existing_ids = set()

    if not os.path.exists(output_path):
        return existing_ids

    fmt = fmt.lower()

    # Determine the dedup column name from the configured branch label.
    # BRANCHES is already populated by the time this is called.
    if BRANCHES:
        dedup_col = f"{next(iter(BRANCHES))} Commit Hash"
    else:
        dedup_col = "Commit Hash"

    if fmt == "csv":
        try:
            with open(output_path, "r", newline="", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                fieldnames = reader.fieldnames or []
                # Find any column ending in ' Commit Hash' (handles label changes)
                hash_col = next(
                    (fn for fn in fieldnames if fn.endswith(" Commit Hash")
                     and not fn.startswith("Fix-")),
                    None
                )
                if not hash_col:
                    print(f"[WARN] Existing CSV has no '<label> Commit Hash' column — treating as empty.")
                    return existing_ids
                for row in reader:
                    cid = (row.get(hash_col) or "").strip()
                    if cid:
                        existing_ids.add(cid)
            print(f"[*] Existing CSV loaded: {output_path}  ({len(existing_ids)} commits already recorded)")
        except Exception as e:
            print(f"[WARN] Could not read existing CSV '{output_path}': {e}  — starting fresh.")

    elif fmt == "xlsx":
        try:
            import openpyxl
            wb = openpyxl.load_workbook(output_path, read_only=True, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            hash_col_idx = next(
                (i for i, h in enumerate(headers)
                 if h and str(h).endswith(" Commit Hash") and not str(h).startswith("Fix-")),
                None
            )
            if hash_col_idx is None:
                print(f"[WARN] Existing XLSX has no '<label> Commit Hash' column — treating as empty.")
                return existing_ids
            for row in ws.iter_rows(min_row=2, values_only=True):
                cid = (row[hash_col_idx] or "")
                if cid:
                    existing_ids.add(str(cid).strip())
            wb.close()
            print(f"[*] Existing XLSX loaded: {output_path}  ({len(existing_ids)} commits already recorded)")
        except ImportError:
            print("[WARN] openpyxl not installed — cannot read existing XLSX. Starting fresh.")
        except Exception as e:
            print(f"[WARN] Could not read existing XLSX '{output_path}': {e}  — starting fresh.")

    elif fmt == "html":
        print(f"[WARN] Deduplication is not supported for HTML output — "
              f"existing file will be rewritten with all rows.")

    return existing_ids


def load_existing_rows_csv(output_path: str) -> list:
    """
    Read all existing rows from a CSV file and return them as a list of dicts.
    Used when appending new rows so the full file can be rewritten with a
    unified, up-to-date fieldnames header.
    Returns an empty list if the file does not exist or cannot be read.
    """
    if not os.path.exists(output_path):
        return []
    try:
        with open(output_path, "r", newline="", encoding="utf-8") as f:
            return list(csv.DictReader(f))
    except Exception as e:
        print(f"[WARN] Could not re-read existing CSV for append: {e}")
        return []


def load_existing_rows_xlsx(output_path: str) -> list:
    """
    Read all existing rows from an XLSX file and return them as a list of dicts.
    Returns an empty list if the file does not exist or cannot be read.
    """
    if not os.path.exists(output_path):
        return []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(output_path, read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i]: (row[i] or "") for i in range(len(headers))})
        wb.close()
        return rows
    except Exception as e:
        print(f"[WARN] Could not re-read existing XLSX for append: {e}")
        return []


def write_csv(rows: list, output_path: str):
    """Write rows to a plain CSV file (always rewrites with full unified header)."""
    fieldnames = build_fieldnames(rows)
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            complete = {fn: row.get(fn, "") for fn in fieldnames}
            writer.writerow(complete)
    print(f"\n[✓] CSV written → {output_path}")
    print(f"    Columns : {len(fieldnames)}")
    print(f"    Rows    : {len(rows)}")


def write_html(rows: list, output_path: str):
    """Write rows to a colour-coded, filterable HTML table."""
    fieldnames = build_fieldnames(rows)

    html_rows = []
    for row in rows:
        sev   = row.get("Severity", "")
        app   = row.get("Applicability", "")
        cells = []
        for fn in fieldnames:
            val = row.get(fn, "") or ""
            # Hyperlink cells that look like URLs
            if str(val).startswith("http"):
                cell = f'<td><a href="{val}" target="_blank">{val}</a></td>'
            else:
                cell = f"<td>{val}</td>"
            cells.append(cell)

        sev_bg = _SEVERITY_COLOURS.get(sev, "")
        app_bg = _APPLICABILITY_COLOURS.get(app, "")
        # Row style: severity colour if present, else applicability colour
        bg = sev_bg or app_bg
        style = f' style="background:{bg}"' if bg else ""
        html_rows.append(f"  <tr{style}>{''.join(cells)}</tr>")

    headers = "".join(f"<th>{h}</th>" for h in fieldnames)

    rows_html = "\n".join(html_rows)
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <title>Kernel Commit Analysis</title>
  <style>
    body  {{ font-family: Arial, sans-serif; font-size: 12px; margin: 16px; }}
    table {{ border-collapse: collapse; width: 100%; }}
    th, td {{ border: 1px solid #ccc; padding: 4px 8px; white-space: nowrap; }}
    th    {{ background: #2c3e50; color: #fff; position: sticky; top: 0; cursor: pointer; }}
    tr:hover td {{ filter: brightness(92%); }}
    input {{ margin-bottom: 10px; padding: 6px; width: 320px; font-size: 13px; }}
  </style>
  <script>
    function filterTable() {{
      var q = document.getElementById('search').value.toLowerCase();
      var rows = document.querySelectorAll('#tbl tbody tr');
      rows.forEach(function(r) {{
        r.style.display = r.innerText.toLowerCase().includes(q) ? '' : 'none';
      }});
    }}
    function sortTable(n) {{
      var tbl = document.getElementById('tbl'), d=1;
      var rows = Array.from(tbl.querySelectorAll('tbody tr'));
      if (tbl._lastSort===n) {{ d = -tbl._lastDir; }} else {{ d=1; }}
      tbl._lastSort=n; tbl._lastDir=d;
      rows.sort(function(a,b){{
        var x = a.cells[n].innerText.toLowerCase();
        var y = b.cells[n].innerText.toLowerCase();
        return x > y ? d : x < y ? -d : 0;
      }});
      rows.forEach(function(r){{ tbl.querySelector('tbody').appendChild(r); }});
    }}
  </script>
</head>
<body>
  <h2>Kernel Commit Analysis — {len(rows)} commits</h2>
  <input id="search" onkeyup="filterTable()" placeholder="Filter rows...">
  <table id="tbl">
    <thead><tr>{''.join(f'<th onclick="sortTable({i})">{h}</th>' for i,h in enumerate(fieldnames))}</tr></thead>
    <tbody>
{rows_html}
    </tbody>
  </table>
</body>
</html>"""

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"\n[✓] HTML written → {output_path}")
    print(f"    Columns : {len(fieldnames)}")
    print(f"    Rows    : {len(rows)}")


def write_xlsx(rows: list, output_path: str):
    """Write rows to a colour-coded Excel (.xlsx) file using openpyxl."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[ERROR] openpyxl is required for --format xlsx.")
        print("        Install it with:  pip install openpyxl")
        sys.exit(1)

    fieldnames = build_fieldnames(rows)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kernel Commits"

    # ── Header row ────────────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="2C3E50")
    header_font = Font(bold=True, color="FFFFFF")
    thin_side   = Side(style="thin", color="999999")
    thin_border = Border(left=thin_side, right=thin_side,
                         top=thin_side, bottom=thin_side)

    for col_idx, header in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    ws.freeze_panes = "A2"   # freeze header row

    # ── Data rows ─────────────────────────────────────────────────────────────
    def hex_to_argb(hex_colour: str) -> str:
        """Convert '#rrggbb' → 'FFrrggbb' for openpyxl."""
        return "FF" + hex_colour.lstrip("#").upper()

    for row_idx, row in enumerate(rows, 2):
        sev = row.get("Severity", "")
        app = row.get("Applicability", "")
        bg_hex = _SEVERITY_COLOURS.get(sev) or _APPLICABILITY_COLOURS.get(app)
        row_fill = PatternFill("solid", fgColor=hex_to_argb(bg_hex)) if bg_hex else None

        for col_idx, fn in enumerate(fieldnames, 1):
            val  = row.get(fn, "") or ""
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border    = thin_border
            cell.alignment = Alignment(wrap_text=False)
            if row_fill:
                cell.fill = row_fill

    # ── Auto-fit column widths (capped at 60) ─────────────────────────────────
    for col_idx, fn in enumerate(fieldnames, 1):
        max_len = max(
            len(str(fn)),
            *(len(str(row.get(fn, "") or "")) for row in rows),
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    # ── Auto-filter on header row ─────────────────────────────────────────────
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    print(f"\n[✓] XLSX written → {output_path}")
    print(f"    Columns : {len(fieldnames)}")
    print(f"    Rows    : {len(rows)}")


def _detect_branch_changes(existing_rows: list) -> None:
    """
    Compare branch columns found in existing_rows against the currently
    configured BRANCHES dict.  Prints clear warnings for:
      - branches present in the file but NOT in current config (removed)
      - branches in current config but NOT in the file (newly added)
    Old rows will have blank cells for any newly added branch columns.
    This is purely informational — processing continues either way.
    """
    if not existing_rows:
        return

    # Collect branch labels that actually exist as columns in the old file.
    # A branch labelled "X.Y.z" produces columns like "X.Y.z Commit Hash".
    # We detect them by looking for keys ending in " Commit Hash".
    file_branches = set()
    for key in existing_rows[0].keys():
        if key.endswith(" Commit Hash"):
            label = key[: -len(" Commit Hash")]
            # Exclude Fix-N prefixed columns (e.g. "Fix-1 6.1.y Commit Hash")
            if not label.startswith("Fix-"):
                file_branches.add(label)

    current_branches = set(BRANCHES.keys())

    added   = current_branches - file_branches
    removed = file_branches - current_branches

    if not added and not removed:
        return   # branches are identical — nothing to warn about

    print("\n[WARN] Branch configuration differs from existing output file:")
    if added:
        print(f"  Newly added   (old rows will have BLANK cells for these): {sorted(added)}")
    if removed:
        print(f"  Removed       (columns still present in file, just not updated): {sorted(removed)}")
    print("  To backfill old rows with new branch data, re-run without --output\n"
          "  pointing at the existing file, then merge manually.\n")


def write_output(new_rows: list, output_path: str, fmt: str):
    """
    Merge new_rows with any existing rows in output_path, then write everything.
    Dispatches to the correct writer based on fmt ('csv', 'html', 'xlsx').
    """
    fmt = fmt.lower()
    # (Extension already normalised in main() before this call)

    # Load pre-existing rows so we can combine old + new into one file
    if fmt == "csv":
        existing_rows = load_existing_rows_csv(output_path)
    elif fmt == "xlsx":
        existing_rows = load_existing_rows_xlsx(output_path)
    else:
        existing_rows = []   # HTML: always rewritten (no structured append)

    if existing_rows:
        # Warn if branch config changed since the file was last written
        _detect_branch_changes(existing_rows)
        print(f"[*] Merging {len(existing_rows)} existing row(s) + {len(new_rows)} new row(s)")

    all_rows = existing_rows + new_rows

    if fmt == "csv":
        write_csv(all_rows, output_path)
    elif fmt == "html":
        write_html(all_rows, output_path)
    elif fmt == "xlsx":
        write_xlsx(all_rows, output_path)
    else:
        print(f"[ERROR] Unknown format: {fmt!r}. Choose csv, html, or xlsx.")
        sys.exit(1)


# ═══════════════════════════════════════════════════════════════════════════════
#  Entry point
# ═══════════════════════════════════════════════════════════════════════════════

# Thread-safe lock used by the progress bar when running in parallel
_pbar_lock = threading.Lock()


def _process_one(args_tuple):
    """
    Worker function for parallel execution.
    Unpacks a tuple so it can be passed via ThreadPoolExecutor.map().
    Returns (original_index, row_dict) to allow result re-ordering.
    """
    idx, commit_id, stable_hash, repo_root, debug, pbar = args_tuple
    row = process_commit(commit_id, repo_root,
                         debug=debug, stable_hash=stable_hash)
    with _pbar_lock:
        pbar.update(1)
    return idx, row


def main():
    parser = argparse.ArgumentParser(
        description=(
            "Generate a kernel commit analysis report (CSV / HTML / XLSX) "
            "from a stable-tree log file.  The log file format is:\n"
            "  <version>  <stable_hash>  <description>\n"
            "e.g.:  6.1.167  4ec349af3ef7...  selftests: net: ...\n\n"
            "Mainline commit IDs are resolved automatically from each stable "
            "commit's body text (no git log --grep needed).  First-Version is "
            "taken directly from the log file (no git tag --contains needed)."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--file", "-f",
        required=True,
        help=(
            "Path to the stable-tree log file.  Each line must be:\n"
            "  <version>  <stable_hash>  <description>\n"
            "Blank lines and lines starting with '#' are ignored."
        ),
    )
    parser.add_argument(
        "--branch", "-b",
        required=True,
        dest="branch",
        help=(
            "Single branch spec: 'label:remote/branch[:config_path]'\n"
            "  label        — human name used in column headers, e.g. 6.1.y\n"
            "  remote/branch — git ref,  e.g. origin/linux-6.1.y\n"
            "  config_path  — optional path to kernel .config for this branch\n"
            "Example: '6.1.y:origin/linux-6.1.y:config-6.1.123'"
        ),
    )
    parser.add_argument(
        "--output", "-o",
        default=OUTPUT_DEFAULT,
        help=f"Output file path (default: {OUTPUT_DEFAULT}). "
             "Extension is auto-adjusted to match --format.",
    )
    parser.add_argument(
        "--format",
        default="csv",
        dest="fmt",
        choices=["csv", "html", "xlsx"],
        help="Output format: csv (default), html, or xlsx.",
    )
    parser.add_argument(
        "--repo", "-r",
        default=".",
        help="Path to the Linux git repo (default: current directory).",
    )
    parser.add_argument(
        "--parallel",
        action="store_true",
        default=False,
        help=(
            "Enable parallel commit processing using cpu_count // 2 workers. "
            "Speeds up large batches. Without this flag, commits are processed "
            "one at a time (safer for debugging)."
        ),
    )
    parser.add_argument(
        "--debug",
        action="store_true",
        default=False,
        help="Print step-by-step applicability trace for every commit.",
    )

    args = parser.parse_args()

    # ── Resolve ALL user-supplied paths to absolute BEFORE chdir ─────────────
    repo_path   = os.path.abspath(os.path.expanduser(args.repo))
    output_path = os.path.abspath(os.path.expanduser(args.output))
    input_file  = os.path.abspath(os.path.expanduser(args.file))

    # Normalise output extension to match --format
    _fmt_now = args.fmt.lower()
    _expected_ext = {"csv": ".csv", "html": ".html", "xlsx": ".xlsx"}[_fmt_now]
    _base, _ext = os.path.splitext(output_path)
    if _ext.lower() != _expected_ext:
        output_path = _base + _expected_ext
        print(f"[*] Output path adjusted to match format: {output_path}")

    # ── Parse --branch spec ───────────────────────────────────────────────────
    global BRANCHES
    spec  = args.branch.strip()
    parts = spec.split(":")
    # Accept:  label:remote/branch
    #          label:remote/branch:config_path
    if len(parts) < 2:
        print(f"[ERROR] --branch must be 'label:remote/branch[:config_path]', got: {spec!r}")
        sys.exit(1)
    b_label    = parts[0].strip()
    b_ref      = parts[1].strip()
    b_cfg_path = (os.path.abspath(os.path.expanduser(parts[2].strip()))
                  if len(parts) >= 3 and parts[2].strip() else None)
    BRANCHES[b_label] = {"ref": b_ref, "cfg_path": b_cfg_path, "kernel_cfg": {}}
    print(f"[*] Branch configured: {b_label}  →  {b_ref}"
          + (f"  (config: {b_cfg_path})" if b_cfg_path else "  (no config)"))

    # ── Validate repo path and chdir ──────────────────────────────────────────
    if not os.path.isdir(repo_path):
        print(f"[ERROR] Repo path not found: {repo_path}")
        sys.exit(1)
    os.chdir(repo_path)
    repo_root = os.getcwd()
    print(f"[*] Working in repo: {repo_root}")

    # ── Load kernel config for the branch ────────────────────────────────────
    if b_cfg_path:
        BRANCHES[b_label]["kernel_cfg"] = load_kernel_config(b_cfg_path)
    else:
        BRANCHES[b_label]["kernel_cfg"] = {}

    # ── Parse the stable-tree log file ───────────────────────────────────────
    if not os.path.isfile(input_file):
        print(f"[ERROR] Log file not found: {input_file}")
        sys.exit(1)

    raw_log = parse_log_file(input_file)   # stable_hash → entry dict

    # ── Populate STABLE_LOG_MAP (keyed by stable hash for fast lookup) ────────
    # collect_commit_info() looks up version directly from this map —
    # no git tag --contains needed.
    global STABLE_LOG_MAP
    STABLE_LOG_MAP = raw_log   # already keyed by stable_hash

    # work list: just the stable hashes in log order (already deduplicated
    # by parse_log_file since it uses the hash as the dict key)
    stable_hashes = list(raw_log.keys())
    print(f"[*] Stable commits loaded    : {len(stable_hashes)}")

    # ── Skip commits already recorded in the output file ─────────────────────
    already_done = load_existing_commit_ids(output_path, args.fmt)
    if already_done:
        before        = len(stable_hashes)
        stable_hashes = [sh for sh in stable_hashes if sh not in already_done]
        skipped       = before - len(stable_hashes)
        if skipped:
            print(f"[*] Skipping {skipped} commit(s) already in existing output file.")

    if not stable_hashes:
        print("[*] All input commits are already recorded — nothing to do.")
        sys.exit(0)

    total = len(stable_hashes)
    print(f"[*] Total commits to process : {total}")
    print(f"[*] Output format            : {args.fmt.upper()}")

    # ── Process commits ───────────────────────────────────────────────────────
    rows = [None] * total   # pre-allocate to preserve input order

    if args.parallel:
        cpu_count = os.cpu_count() or 1
        n_workers = max(1, cpu_count // 2)
        print(f"[*] Parallel mode            : {n_workers} workers "
              f"(cpu_count={cpu_count} // 2)")

        with make_progress(total, desc="Commits") as pbar:
            work_items = [
                (idx, sh, sh, repo_root, args.debug, pbar)
                for idx, sh in enumerate(stable_hashes)
            ]
            with concurrent.futures.ThreadPoolExecutor(max_workers=n_workers) as executor:
                for idx, row in executor.map(_process_one, work_items):
                    rows[idx] = row

    else:
        print(f"[*] Parallel mode            : off (single-core)")
        with make_progress(total, desc="Commits") as pbar:
            for idx, sh in enumerate(stable_hashes):
                rows[idx] = process_commit(
                    sh, repo_root,
                    debug=args.debug, stable_hash=sh,
                )
                pbar.update(1)

    # ── Write output ──────────────────────────────────────────────────────────
    write_output(rows, output_path, args.fmt)
    print_summary(rows)


if __name__ == "__main__":
    main()
